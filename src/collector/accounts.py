from __future__ import annotations

import csv
from pathlib import Path

import yaml
from openpyxl import load_workbook

PLATFORMS = ("bilibili", "weibo", "douyin", "kuaishou")

_HEADERS = {
    "bilibili": "# 每行一个账号；account_id 是 B 站 mid",
    "weibo": "# account_id 是数字 uid，从 weibo.com/u/<uid> 地址栏复制",
    "douyin": "# account_id 是抖音 sec_uid，从 douyin.com/user/<sec_uid> 地址栏复制",
    "kuaishou": "# account_id 是快手个人主页 URL 末段（kuaishou.com/profile/<id>）",
}


def _config_path(config_root: Path, platform: str) -> Path:
    return config_root / f"{platform}.yaml"


def load_accounts_raw(config_root: Path, platform: str) -> list[dict]:
    """读 configs/<platform>.yaml，返回 [{account_id, account_name}, ...]；缺文件返回 []。"""
    p = _config_path(config_root, platform)
    if not p.exists():
        return []
    data = yaml.safe_load(p.read_text("utf-8")) or []
    # account_id 一律规整成 str（防 yaml 把纯数字读成 int）
    out = []
    for r in data:
        out.append(
            {"account_id": str(r["account_id"]), "account_name": str(r.get("account_name", ""))}
        )
    return out


def write_accounts(config_root: Path, platform: str, accounts: list[dict]) -> None:
    """结构化写回，并在文件头重写该平台的说明注释。"""
    config_root.mkdir(parents=True, exist_ok=True)
    body = yaml.safe_dump(accounts, allow_unicode=True, sort_keys=False)
    _config_path(config_root, platform).write_text(
        _HEADERS[platform] + "\n" + body, encoding="utf-8"
    )


def add_account(config_root: Path, platform: str, account_id: str, account_name: str) -> bool:
    """加一个账号。已存在（同 account_id）返回 False 且不改文件，否则写入返回 True。"""
    accounts = load_accounts_raw(config_root, platform)
    if any(a["account_id"] == str(account_id) for a in accounts):
        return False
    accounts.append({"account_id": str(account_id), "account_name": str(account_name)})
    write_accounts(config_root, platform, accounts)
    return True


def _read_csv(path: Path) -> list[dict]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def _read_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        out.append({header[i]: (r[i] if i < len(r) else None) for i in range(len(header))})
    return out


def parse_import_file(path: Path) -> tuple[list[dict], list[tuple[int, str]]]:
    """解析导入文件，返回 (合法行, 错误行)。
    合法行：[{platform, account_id, account_name}]；account_name 为空时用 account_id 兜底。
    错误行：[(行号, 原因)]，行号含表头（数据从第 2 行起）。全空行忽略，不算错误。
    """
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        raw = _read_csv(path)
    elif suffix == ".xlsx":
        raw = _read_xlsx(path)
    else:
        raise ValueError(f"不支持的导入格式：{suffix}（仅 .csv / .xlsx）")

    valid: list[dict] = []
    errors: list[tuple[int, str]] = []
    for i, r in enumerate(raw, start=2):  # 第 1 行是表头
        platform = str(r.get("platform") or "").strip()
        aid = str(r.get("account_id") or "").strip()
        name = str(r.get("account_name") or "").strip()
        if not platform and not aid and not name:
            continue  # 全空行忽略
        if platform not in PLATFORMS:
            errors.append((i, f"平台名错: {platform!r}"))
            continue
        if not aid:
            errors.append((i, "account_id 为空"))
            continue
        valid.append({"platform": platform, "account_id": aid, "account_name": name or aid})
    return valid, errors


def validate_account_id(platform: str, account_id: str) -> str | None:
    """按平台对 account_id 做轻量格式提示。不符返回警告串，正常返回 None。仅警告不拦截。"""
    aid = str(account_id)
    if platform in ("bilibili", "weibo"):
        if not aid.isdigit():
            return f"{platform} 的 account_id 一般是纯数字，当前 '{aid}' 可能填错"
    elif platform == "douyin":
        if not aid.startswith("MS4w"):
            return f"抖音 sec_uid 一般以 MS4w 开头，当前 '{aid}' 可能填错"
    elif platform == "kuaishou":
        if not aid.startswith("3x"):
            return f"快手主页 id 一般以 3x 开头，当前 '{aid}' 可能填错"
    return None


def remove_account(config_root: Path, platform: str, account_id: str) -> bool:
    """删一个账号。删到了写回返回 True；不存在返回 False。"""
    accounts = load_accounts_raw(config_root, platform)
    kept = [a for a in accounts if a["account_id"] != str(account_id)]
    if len(kept) == len(accounts):
        return False
    write_accounts(config_root, platform, kept)
    return True
