from datetime import datetime, timezone
from pathlib import Path

import openpyxl

from collector.rendering import COLUMNS, render_xlsx
from collector.schemas import Post


def _post() -> Post:
    return Post(
        platform="bilibili",
        post_id="BV1",
        url="https://www.bilibili.com/video/BV1",
        title="某视频",
        caption="简介",
        cover_url="https://i0.hdslb.com/cover.jpg",
        duration_sec=125,
        media_type="video",
        published_at=datetime(2026, 6, 1, 10, 30, 0, tzinfo=timezone.utc),
        like_count=10,
        comment_count=2,
        share_count=1,
        view_count=300,
        author_id="A1",
        author_name="复旦",
        fetched_at=datetime(2026, 6, 8, tzinfo=timezone.utc),
        raw={},
    )


_EXPECTED = [
    "发布者", "页面源码1", "标题", "视频链接", "播放数",
    "发布时间", "时长", "视频封面链接", "视频封面链接_保存位置",
]


def test_bilibili_columns_match_spec():
    names = [c.name for c in COLUMNS["bilibili"]]
    assert names == _EXPECTED


def test_render_bilibili_xlsx(tmp_path: Path):
    out = tmp_path / "bilibili.xlsx"
    render_xlsx(out, platform="bilibili", posts=[_post()])
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers == _EXPECTED
    row2 = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    assert row2[0] == "复旦"
    assert row2[1] is None or row2[1] == ""
    assert row2[2] == "某视频"
    assert row2[3] == "https://www.bilibili.com/video/BV1"
    assert row2[4] == 300
    assert row2[6] == "00:02:05"
    assert row2[8] is None or row2[8] == ""


def test_douyin_columns_match_spec():
    names = [c.name for c in COLUMNS["douyin"]]
    assert names == [
        "博主名称", "博主简介", "视频标题", "视频链接", "视频点赞数",
        "封面图url", "是否置顶", "发布时间", "视频时长", "评论数",
        "收藏数", "转发数", "页面网址",
    ]
